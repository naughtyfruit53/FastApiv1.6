# app/api/v1/bom.py

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from sqlalchemy import select, delete
from typing import List, Optional
from datetime import datetime
from pydantic import BaseModel, field_validator
import logging
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from fastapi.responses import StreamingResponse

from app.core.database import get_db
from app.core.security import get_current_user
from app.api.v1.auth import get_current_active_user
from app.models.vouchers import BillOfMaterials, BOMComponent
from app.models.vouchers.purchase import PurchaseOrderItem
from app.models import Product
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import joinedload
from app.schemas.base import ProductResponse  # Added import for ProductResponse

logger = logging.getLogger(__name__)
router = APIRouter()


class BOMComponentCreate(BaseModel):
    component_item_id: int
    quantity_required: float
    unit: str
    unit_cost: Optional[float] = 0.0
    wastage_percentage: Optional[float] = 0.0
    is_optional: Optional[bool] = False
    sequence: Optional[int] = 0
    notes: Optional[str] = None


class BOMComponentResponse(BOMComponentCreate):
    id: int
    bom_id: int
    organization_id: int
    total_cost: float
    component_item: Optional[ProductResponse] = None  # Changed from dict to ProductResponse

    class Config:
        from_attributes = True


class BOMCreate(BaseModel):
    bom_name: str
    output_item_id: int
    output_quantity: Optional[float] = 1.0
    version: Optional[str] = "1.0"
    validity_start: Optional[datetime] = None
    validity_end: Optional[datetime] = None
    description: Optional[str] = None
    notes: Optional[str] = None
    material_cost: Optional[float] = 0.0
    labor_cost: Optional[float] = 0.0
    overhead_cost: Optional[float] = 0.0
    components: List[BOMComponentCreate] = []

    @field_validator('validity_start', 'validity_end', mode='before')
    @classmethod
    def handle_empty_datetime(cls, v):
        if isinstance(v, str) and v.strip() == '':
            return None
        return v


class BOMUpdate(BaseModel):
    bom_name: Optional[str] = None
    output_item_id: Optional[int] = None
    output_quantity: Optional[float] = None
    version: Optional[str] = None
    validity_start: Optional[datetime] = None
    validity_end: Optional[datetime] = None
    description: Optional[str] = None
    notes: Optional[str] = None
    material_cost: Optional[float] = None
    labor_cost: Optional[float] = None
    overhead_cost: Optional[float] = None
    is_active: Optional[bool] = None
    components: Optional[List[BOMComponentCreate]] = None

    @field_validator('validity_start', 'validity_end', mode='before')
    @classmethod
    def handle_empty_datetime(cls, v):
        if isinstance(v, str) and v.strip() == '':
            return None
        return v


class BOMResponse(BaseModel):
    id: int
    organization_id: int
    bom_name: str
    output_item_id: int
    output_quantity: float
    version: str
    validity_start: Optional[datetime] = None
    validity_end: Optional[datetime] = None
    is_active: bool
    description: Optional[str] = None
    notes: Optional[str] = None
    material_cost: float
    labor_cost: float
    overhead_cost: float
    total_cost: float
    created_at: datetime
    updated_at: Optional[datetime] = None
    output_item: Optional[ProductResponse] = None  # Changed from dict to ProductResponse
    components: List[BOMComponentResponse] = []

    class Config:
        from_attributes = True


@router.get("/", response_model=List[BOMResponse])
async def get_boms(
    skip: int = 0,
    limit: int = 100,
    search: Optional[str] = None,
    is_active: Optional[bool] = None,
    output_item_id: Optional[int] = None,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Get all BOMs for the organization"""
    stmt = select(BillOfMaterials).options(
        joinedload(BillOfMaterials.output_item).joinedload(Product.files),
        joinedload(BillOfMaterials.components).joinedload(BOMComponent.component_item).joinedload(Product.files)
    ).where(
        BillOfMaterials.organization_id == current_user.organization_id
    )
    
    if search:
        stmt = stmt.where(
            BillOfMaterials.bom_name.ilike(f"%{search}%")
        )
    
    if is_active is not None:
        stmt = stmt.where(BillOfMaterials.is_active == is_active)
    
    if output_item_id:
        stmt = stmt.where(BillOfMaterials.output_item_id == output_item_id)
    
    stmt = stmt.offset(skip).limit(limit)
    result = await db.execute(stmt)
    boms = result.unique().scalars().all()
    return boms


@router.get("/{bom_id}", response_model=BOMResponse)
async def get_bom(
    bom_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Get a specific BOM by ID"""
    stmt = select(BillOfMaterials).options(
        joinedload(BillOfMaterials.output_item).joinedload(Product.files),
        joinedload(BillOfMaterials.components).joinedload(BOMComponent.component_item).joinedload(Product.files)
    ).where(
        BillOfMaterials.id == bom_id,
        BillOfMaterials.organization_id == current_user.organization_id
    )
    result = await db.execute(stmt)
    bom = result.unique().scalar_one_or_none()
    
    if not bom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BOM not found"
        )
    
    return bom


@router.post("/", response_model=BOMResponse)
async def create_bom(
    bom_data: BOMCreate,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Create a new BOM"""
    
    # Check if output item exists
    stmt = select(Product).where(
        Product.id == bom_data.output_item_id,
        Product.organization_id == current_user.organization_id
    )
    result = await db.execute(stmt)
    output_item = result.scalar_one_or_none()
    
    if not output_item:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Output item not found"
        )
    
    # Check for duplicate BOM name + version
    stmt = select(BillOfMaterials).where(
        BillOfMaterials.organization_id == current_user.organization_id,
        BillOfMaterials.bom_name == bom_data.bom_name,
        BillOfMaterials.version == bom_data.version
    )
    result = await db.execute(stmt)
    existing_bom = result.scalar_one_or_none()
    
    if existing_bom:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"BOM '{bom_data.bom_name}' version '{bom_data.version}' already exists"
        )
    
    # Calculate total cost
    material_cost = sum(
        comp.quantity_required * comp.unit_cost * (1 + comp.wastage_percentage / 100)
        for comp in bom_data.components
    )
    total_cost = material_cost + bom_data.labor_cost + bom_data.overhead_cost
    
    # Create BOM
    db_bom = BillOfMaterials(
        organization_id=current_user.organization_id,
        bom_name=bom_data.bom_name,
        output_item_id=bom_data.output_item_id,
        output_quantity=bom_data.output_quantity,
        version=bom_data.version,
        validity_start=bom_data.validity_start,
        validity_end=bom_data.validity_end,
        description=bom_data.description,
        notes=bom_data.notes,
        material_cost=material_cost,
        labor_cost=bom_data.labor_cost,
        overhead_cost=bom_data.overhead_cost,
        total_cost=total_cost,
        created_by=current_user.id
    )
    
    db.add(db_bom)
    await db.flush()  # Get the ID
    
    # Create components
    for idx, comp_data in enumerate(bom_data.components):
        # Verify component item exists
        stmt = select(Product).where(
            Product.id == comp_data.component_item_id,
            Product.organization_id == current_user.organization_id
        )
        result = await db.execute(stmt)
        component_item = result.scalar_one_or_none()
        
        if not component_item:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail=f"Component item ID {comp_data.component_item_id} not found"
            )
        
        # Use discounted price if available
        stmt_discount = select(PurchaseOrderItem.discounted_price).where(
            PurchaseOrderItem.product_id == comp_data.component_item_id
        ).order_by(PurchaseOrderItem.id.desc()).limit(1)
        result_discount = await db.execute(stmt_discount)
        discounted_price = result_discount.scalar_one_or_none()
        
        unit_cost = discounted_price if discounted_price else comp_data.unit_cost
        
        total_comp_cost = comp_data.quantity_required * unit_cost * (1 + comp_data.wastage_percentage / 100)
        
        component = BOMComponent(
            organization_id=current_user.organization_id,
            bom_id=db_bom.id,
            component_item_id=comp_data.component_item_id,
            quantity_required=comp_data.quantity_required,
            unit=comp_data.unit,
            unit_cost=unit_cost,
            total_cost=total_comp_cost,
            wastage_percentage=comp_data.wastage_percentage,
            is_optional=comp_data.is_optional,
            sequence=comp_data.sequence or idx,
            notes=comp_data.notes
        )
        db.add(component)
    
    # Update output item: set unit price to total BOM cost per unit and mark as manufactured
    unit_price = total_cost / bom_data.output_quantity if bom_data.output_quantity > 0 else total_cost
    output_item.unit_price = unit_price
    output_item.is_manufactured = True
    
    await db.commit()
    
    # Refresh BOM with relationships loaded
    stmt = select(BillOfMaterials).options(
        joinedload(BillOfMaterials.output_item).joinedload(Product.files),
        joinedload(BillOfMaterials.components).joinedload(BOMComponent.component_item).joinedload(Product.files)
    ).where(BillOfMaterials.id == db_bom.id)
    result = await db.execute(stmt)
    db_bom = result.unique().scalar_one()
    
    return db_bom


@router.put("/{bom_id}", response_model=BOMResponse)
async def update_bom(
    bom_id: int,
    bom_data: BOMUpdate,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Update an existing BOM"""
    
    # Get existing BOM
    stmt = select(BillOfMaterials).where(
        BillOfMaterials.id == bom_id,
        BillOfMaterials.organization_id == current_user.organization_id
    )
    result = await db.execute(stmt)
    db_bom = result.scalar_one_or_none()
    
    if not db_bom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BOM not found"
        )
    
    # Check if BOM is in use (has manufacturing orders)
    from app.models.vouchers import ManufacturingOrder
    stmt = select(ManufacturingOrder).where(
        ManufacturingOrder.bom_id == bom_id,
        ManufacturingOrder.organization_id == current_user.organization_id,
        ManufacturingOrder.production_status.in_(["planned", "in_progress"])
    )
    result = await db.execute(stmt)
    in_use = result.scalar_one_or_none()
    
    if in_use and bom_data.components is not None:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot modify components of BOM that is currently in use in manufacturing orders"
        )
    
    # Update BOM fields
    update_data = bom_data.dict(exclude_unset=True)
    components_data = update_data.pop('components', None)
    
    for field, value in update_data.items():
        setattr(db_bom, field, value)
    
    # Update components if provided
    if components_data is not None:
        # Delete existing components
        await db.execute(delete(BOMComponent).where(BOMComponent.bom_id == bom_id))
        
        # Create new components
        material_cost = 0
        for idx, comp_data in enumerate(components_data):
            # Verify component item exists
            stmt = select(Product).where(
                Product.id == comp_data.component_item_id,
                Product.organization_id == current_user.organization_id
            )
            result = await db.execute(stmt)
            component_item = result.scalar_one_or_none()
            
            if not component_item:
                raise HTTPException(
                    status_code=status.HTTP_404_NOT_FOUND,
                    detail=f"Component item ID {comp_data.component_item_id} not found"
                )
            
            # Use discounted price if available
            stmt_discount = select(PurchaseOrderItem.discounted_price).where(
                PurchaseOrderItem.product_id == comp_data.component_item_id
            ).order_by(PurchaseOrderItem.id.desc()).limit(1)
            result_discount = await db.execute(stmt_discount)
            discounted_price = result_discount.scalar_one_or_none()
            
            unit_cost = discounted_price if discounted_price else comp_data.unit_cost
            
            total_comp_cost = comp_data.quantity_required * unit_cost * (1 + comp_data.wastage_percentage / 100)
            material_cost += total_comp_cost
            
            component = BOMComponent(
                organization_id=current_user.organization_id,
                bom_id=db_bom.id,
                component_item_id=comp_data.component_item_id,
                quantity_required=comp_data.quantity_required,
                unit=comp_data.unit,
                unit_cost=unit_cost,
                total_cost=total_comp_cost,
                wastage_percentage=comp_data.wastage_percentage,
                is_optional=comp_data.is_optional,
                sequence=comp_data.sequence or idx,
                notes=comp_data.notes
            )
            db.add(component)
        
        # Update material cost
        db_bom.material_cost = material_cost
        db_bom.total_cost = material_cost + db_bom.labor_cost + db_bom.overhead_cost
    
    await db.commit()
    
    # Refresh BOM with relationships loaded
    stmt = select(BillOfMaterials).options(
        joinedload(BillOfMaterials.output_item).joinedload(Product.files),
        joinedload(BillOfMaterials.components).joinedload(BOMComponent.component_item).joinedload(Product.files)
    ).where(BillOfMaterials.id == db_bom.id)
    result = await db.execute(stmt)
    db_bom = result.unique().scalar_one()
    
    return db_bom


@router.delete("/{bom_id}")
async def delete_bom(
    bom_id: int,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Delete a BOM (only if not in use)"""
    
    # Get existing BOM
    stmt = select(BillOfMaterials).where(
        BillOfMaterials.id == bom_id,
        BillOfMaterials.organization_id == current_user.organization_id
    )
    result = await db.execute(stmt)
    db_bom = result.scalar_one_or_none()
    
    if not db_bom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BOM not found"
        )
    
    # Check if BOM is in use (has manufacturing orders)
    from app.models.vouchers import ManufacturingOrder
    stmt = select(ManufacturingOrder).where(
        ManufacturingOrder.bom_id == bom_id,
        ManufacturingOrder.organization_id == current_user.organization_id
    )
    result = await db.execute(stmt)
    in_use = result.scalar_one_or_none()
    
    if in_use:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete BOM that is in use in manufacturing orders"
        )
    
    # Delete a BOM (only if not in use)
    await db.delete(db_bom)
    await db.commit()
    
    return {"message": "BOM deleted successfully"}


@router.get("/{bom_id}/cost-breakdown")
async def get_bom_cost_breakdown(
    bom_id: int,
    production_quantity: float = 1.0,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Get detailed cost breakdown for a BOM"""
    
    stmt = select(BillOfMaterials).where(
        BillOfMaterials.id == bom_id,
        BillOfMaterials.organization_id == current_user.organization_id
    )
    result = await db.execute(stmt)
    bom = result.scalar_one_or_none()
    
    if not bom:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="BOM not found"
        )
    
    # Calculate costs for the requested production quantity
    multiplier = production_quantity / bom.output_quantity
    
    component_costs = []
    total_material_cost = 0
    
    for component in bom.components:
        required_qty = component.quantity_required * multiplier
        wastage_qty = required_qty * (component.wastage_percentage / 100)
        total_qty = required_qty + wastage_qty
        component_cost = total_qty * component.unit_cost
        
        component_costs.append({
            "component_name": component.component_item.product_name if component.component_item else "Unknown",
            "required_quantity": required_qty,
            "wastage_quantity": wastage_qty,
            "total_quantity": total_qty,
            "unit_cost": component.unit_cost,
            "total_cost": component_cost,
            "unit": component.unit
        })
        
        total_material_cost += component_cost
    
    return {
        "bom_name": bom.bom_name,
        "output_item": bom.output_item.product_name if bom.output_item else "Unknown",
        "production_quantity": production_quantity,
        "cost_breakdown": {
            "material_cost": total_material_cost,
            "labor_cost": bom.labor_cost * multiplier,
            "overhead_cost": bom.overhead_cost * multiplier,
            "total_cost": (total_material_cost + bom.labor_cost + bom.overhead_cost) * multiplier
        },
        "component_costs": component_costs
    }


@router.get("/export/template")
async def download_bom_template(
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Download BOM import template Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Import Template"
    
    # Define headers
    headers = [
        "BOM Name", "Output Item Code/Name", "Output Quantity", "Version",
        "Description", "Material Cost", "Labor Cost", "Overhead Cost",
        "Component Item Code/Name", "Quantity Required", "Unit", "Unit Cost",
        "Wastage %", "Is Optional", "Sequence", "Notes"
    ]
    
    # Add headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Add sample data row
    sample_data = [
        "Sample BOM", "PROD-001", "1", "1.0", "Sample description",
        "1000", "500", "200", "COMP-001", "2", "pcs", "100",
        "5", "FALSE", "1", "Sample component"
    ]
    for col_num, value in enumerate(sample_data, 1):
        ws.cell(row=2, column=col_num, value=value)
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="BOM_Import_Template.xlsx"'
    }
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


@router.post("/import")
async def import_boms_from_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Import BOMs from Excel file"""
    import pandas as pd
    import io
    
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Only Excel files (.xlsx, .xls) are supported"
        )
    
    try:
        # Read Excel file
        content = await file.read()
        df = pd.read_excel(io.BytesIO(content))
        
        # Clean column names
        df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
        
        # Group by BOM to handle multiple components
        imported_count = 0
        errors = []
        
        bom_groups = df.groupby('bom_name')
        
        for bom_name, bom_data in bom_groups:
            try:
                # Get BOM header info from first row
                first_row = bom_data.iloc[0]
                
                # Find or create output item
                output_item_ref = str(first_row.get('output_item_code/name', '')).strip()
                stmt = select(Product).where(
                    Product.organization_id == current_user.organization_id,
                    (Product.product_code == output_item_ref) | 
                    (Product.product_name == output_item_ref)
                )
                result = await db.execute(stmt)
                output_item = result.scalar_one_or_none()
                
                if not output_item:
                    errors.append(f"Output item '{output_item_ref}' not found for BOM '{bom_name}'")
                    continue
                
                # Check if BOM already exists
                stmt = select(BillOfMaterials).where(
                    BillOfMaterials.organization_id == current_user.organization_id,
                    BillOfMaterials.bom_name == bom_name,
                    BillOfMaterials.version == str(first_row.get('version', '1.0'))
                )
                result = await db.execute(stmt)
                existing_bom = result.scalar_one_or_none()
                
                if existing_bom:
                    errors.append(f"BOM '{bom_name}' version '{first_row.get('version', '1.0')}' already exists")
                    continue
                
                # Create BOM
                material_cost = float(first_row.get('material_cost', 0) or 0)
                labor_cost = float(first_row.get('labor_cost', 0) or 0)
                overhead_cost = float(first_row.get('overhead_cost', 0) or 0)
                
                new_bom = BillOfMaterials(
                    organization_id=current_user.organization_id,
                    bom_name=bom_name,
                    output_item_id=output_item.id,
                    output_quantity=float(first_row.get('output_quantity', 1) or 1),
                    version=str(first_row.get('version', '1.0')),
                    description=str(first_row.get('description', '') or ''),
                    material_cost=material_cost,
                    labor_cost=labor_cost,
                    overhead_cost=overhead_cost,
                    total_cost=material_cost + labor_cost + overhead_cost,
                    created_by=current_user.id
                )
                
                db.add(new_bom)
                await db.flush()
                
                # Add components
                for idx, row in bom_data.iterrows():
                    component_ref = str(row.get('component_item_code/name', '')).strip()
                    if not component_ref:
                        continue
                    
                    stmt = select(Product).where(
                        Product.organization_id == current_user.organization_id,
                        (Product.product_code == component_ref) | 
                        (Product.product_name == component_ref)
                    )
                    result = await db.execute(stmt)
                    component_item = result.scalar_one_or_none()
                    
                    if not component_item:
                        errors.append(f"Component '{component_ref}' not found for BOM '{bom_name}'")
                        continue
                    
                    qty_required = float(row.get('quantity_required', 0) or 0)
                    unit_cost = float(row.get('unit_cost', 0) or 0)
                    wastage = float(row.get('wastage_%', 0) or 0)
                    total_cost = qty_required * unit_cost * (1 + wastage / 100)
                    
                    component = BOMComponent(
                        organization_id=current_user.organization_id,
                        bom_id=new_bom.id,
                        component_item_id=component_item.id,
                        quantity_required=qty_required,
                        unit=str(row.get('unit', 'pcs')),
                        unit_cost=unit_cost,
                        total_cost=total_cost,
                        wastage_percentage=wastage,
                        is_optional=str(row.get('is_optional', 'FALSE')).upper() == 'TRUE',
                        sequence=int(row.get('sequence', idx) or idx),
                        notes=str(row.get('notes', '') or '')
                    )
                    db.add(component)
                
                imported_count += 1
                
            except Exception as e:
                errors.append(f"Error importing BOM '{bom_name}': {str(e)}")
                continue
        
        await db.commit()
        
        return {
            "message": f"Successfully imported {imported_count} BOMs",
            "imported_count": imported_count,
            "errors": errors if errors else None
        }
        
    except Exception as e:
        await db.rollback()
        logger.error(f"Error importing BOMs: {e}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to import BOMs: {str(e)}"
        )


@router.get("/export")
async def export_boms_to_excel(
    db: AsyncSession = Depends(get_db),
    current_user = Depends(get_current_active_user)
):
    """Export BOMs to Excel file"""
    
    # Get all BOMs for the organization
    stmt = select(BillOfMaterials).where(
        BillOfMaterials.organization_id == current_user.organization_id
    )
    result = await db.execute(stmt)
    boms = result.scalars().all()
    
    if not boms:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No BOMs found to export"
        )
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM Export"
    
    # Headers
    headers = [
        "BOM Name", "Output Item Code", "Output Item Name", "Output Quantity", 
        "Version", "Description", "Material Cost", "Labor Cost", "Overhead Cost",
        "Total Cost", "Component Item Code", "Component Item Name", 
        "Quantity Required", "Unit", "Unit Cost", "Total Component Cost",
        "Wastage %", "Is Optional", "Sequence", "Notes"
    ]
    
    # Add headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Add data
    row_num = 2
    for bom in boms:
        for component in bom.components:
            ws.cell(row=row_num, column=1, value=bom.bom_name)
            ws.cell(row=row_num, column=2, value=bom.output_item.product_code if bom.output_item else '')
            ws.cell(row=row_num, column=3, value=bom.output_item.product_name if bom.output_item else '')
            ws.cell(row=row_num, column=4, value=bom.output_quantity)
            ws.cell(row=row_num, column=5, value=bom.version)
            ws.cell(row=row_num, column=6, value=bom.description or '')
            ws.cell(row=row_num, column=7, value=bom.material_cost)
            ws.cell(row=row_num, column=8, value=bom.labor_cost)
            ws.cell(row=row_num, column=9, value=bom.overhead_cost)
            ws.cell(row=row_num, column=10, value=bom.total_cost)
            ws.cell(row=row_num, column=11, value=component.component_item.product_code if component.component_item else '')
            ws.cell(row=row_num, column=12, value=component.component_item.product_name if component.component_item else '')
            ws.cell(row=row_num, column=13, value=component.quantity_required)
            ws.cell(row=row_num, column=14, value=component.unit)
            ws.cell(row=row_num, column=15, value=component.unit_cost)
            ws.cell(row=row_num, column=16, value=component.total_cost)
            ws.cell(row=row_num, column=17, value=component.wastage_percentage)
            ws.cell(row=row_num, column=18, value='TRUE' if component.is_optional else 'FALSE')
            ws.cell(row=row_num, column=19, value=component.sequence)
            ws.cell(row=row_num, column=20, value=component.notes or '')
            row_num += 1
    
    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    headers = {
        'Content-Disposition': 'attachment; filename="BOM_Export.xlsx"'
    }
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )