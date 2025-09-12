"""
Financial Export Utilities - Excel and CSV export functionality for financial models and reports
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, BarChart, Reference
from typing import Dict, List, Any, Optional
from io import BytesIO
import json
from datetime import datetime, date
from decimal import Decimal


class FinancialExportService:
    """Service for exporting financial models and forecasts to Excel/CSV"""
    
    def __init__(self):
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.center_alignment = Alignment(horizontal='center', vertical='center')
    
    def export_dcf_model_to_excel(
        self,
        financial_model: Dict[str, Any],
        dcf_model: Dict[str, Any],
        scenarios: Optional[List[Dict[str, Any]]] = None
    ) -> BytesIO:
        """Export DCF model to comprehensive Excel workbook"""
        
        # Create workbook and worksheets
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create worksheets
        summary_ws = wb.create_sheet("Executive Summary")
        assumptions_ws = wb.create_sheet("Assumptions")
        projections_ws = wb.create_sheet("Financial Projections")
        dcf_ws = wb.create_sheet("DCF Valuation")
        sensitivity_ws = wb.create_sheet("Sensitivity Analysis")
        
        if scenarios:
            scenarios_ws = wb.create_sheet("Scenario Analysis")
        
        # Populate Executive Summary
        self._create_executive_summary(summary_ws, financial_model, dcf_model)
        
        # Populate Assumptions
        self._create_assumptions_sheet(assumptions_ws, financial_model['assumptions'])
        
        # Populate Financial Projections
        self._create_projections_sheet(projections_ws, dcf_model['cash_flow_projections'])
        
        # Populate DCF Valuation
        self._create_dcf_sheet(dcf_ws, dcf_model)
        
        # Populate Sensitivity Analysis
        self._create_sensitivity_sheet(sensitivity_ws, financial_model, dcf_model)
        
        # Populate Scenarios if provided
        if scenarios:
            self._create_scenarios_sheet(scenarios_ws, scenarios)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_executive_summary(self, ws, financial_model: Dict, dcf_model: Dict):
        """Create executive summary worksheet"""
        
        # Title
        ws['A1'] = f"Financial Model: {financial_model['model_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Model info
        ws['A3'] = "Model Type:"
        ws['B3'] = financial_model['model_type'].upper()
        ws['A4'] = "Analysis Period:"
        ws['B4'] = f"{financial_model['analysis_start_date']} to {financial_model['analysis_end_date']}"
        ws['A5'] = "Created:"
        ws['B5'] = financial_model['created_at'][:10]
        ws['A6'] = "Version:"
        ws['B6'] = financial_model['model_version']
        
        # Key Results
        ws['A8'] = "KEY VALUATION RESULTS"
        ws['A8'].font = self.header_font
        ws['A8'].fill = self.header_fill
        ws.merge_cells('A8:D8')
        
        results = [
            ["WACC", f"{dcf_model['wacc']:.2f}%"],
            ["Enterprise Value", f"${dcf_model['enterprise_value']:,.0f}"],
            ["Equity Value", f"${dcf_model['equity_value']:,.0f}"],
            ["Value per Share", f"${dcf_model.get('value_per_share', 0):.2f}" if dcf_model.get('value_per_share') else "N/A"],
            ["PV of FCF", f"${dcf_model['pv_of_fcf']:,.0f}"],
            ["PV of Terminal Value", f"${dcf_model['pv_of_terminal_value']:,.0f}"],
        ]
        
        for i, (metric, value) in enumerate(results, start=9):
            ws[f'A{i}'] = metric
            ws[f'B{i}'] = value
            ws[f'A{i}'].font = Font(bold=True)
        
        # Value breakdown
        ws['A16'] = "VALUE BREAKDOWN"
        ws['A16'].font = self.header_font
        ws['A16'].fill = self.header_fill
        ws.merge_cells('A16:C16')
        
        ws['A17'] = "Component"
        ws['B17'] = "Value ($)"
        ws['C17'] = "% of Enterprise Value"
        
        enterprise_value = dcf_model['enterprise_value']
        pv_fcf_pct = (dcf_model['pv_of_fcf'] / enterprise_value) * 100
        pv_terminal_pct = (dcf_model['pv_of_terminal_value'] / enterprise_value) * 100
        
        breakdown = [
            ["PV of Free Cash Flow", f"${dcf_model['pv_of_fcf']:,.0f}", f"{pv_fcf_pct:.1f}%"],
            ["PV of Terminal Value", f"${dcf_model['pv_of_terminal_value']:,.0f}", f"{pv_terminal_pct:.1f}%"],
            ["Enterprise Value", f"${enterprise_value:,.0f}", "100.0%"],
        ]
        
        for i, (component, value, percent) in enumerate(breakdown, start=18):
            ws[f'A{i}'] = component
            ws[f'B{i}'] = value
            ws[f'C{i}'] = percent
            
            if component == "Enterprise Value":
                for col in ['A', 'B', 'C']:
                    ws[f'{col}{i}'].font = Font(bold=True)
        
        # Apply styling
        for row in ws.iter_rows(min_row=17, max_row=20, min_col=1, max_col=3):
            for cell in row:
                cell.border = self.border
                cell.alignment = self.center_alignment
    
    def _create_assumptions_sheet(self, ws, assumptions: Dict):
        """Create assumptions worksheet"""
        
        ws['A1'] = "MODEL ASSUMPTIONS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        # Revenue assumptions
        ws['A3'] = "REVENUE ASSUMPTIONS"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill
        ws.merge_cells('A3:C3')
        
        revenue_data = [
            ["Base Revenue", f"${assumptions.get('base_revenue', 0):,.0f}"],
            ["Gross Margin", f"{assumptions.get('gross_margin', 0):.1f}%"],
            ["Operating Margin", f"{assumptions.get('operating_margin', 0):.1f}%"],
        ]
        
        row = 4
        for assumption, value in revenue_data:
            ws[f'A{row}'] = assumption
            ws[f'B{row}'] = value
            row += 1
        
        # Growth rates
        ws[f'A{row}'] = "Revenue Growth Rates:"
        row += 1
        
        growth_rates = assumptions.get('revenue_growth_rates', [])
        for i, rate in enumerate(growth_rates):
            ws[f'A{row}'] = f"Year {i+1}"
            ws[f'B{row}'] = f"{rate:.1f}%"
            row += 1
        
        row += 1
        
        # Financial assumptions
        ws[f'A{row}'] = "FINANCIAL ASSUMPTIONS"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        financial_data = [
            ["Tax Rate", f"{assumptions.get('tax_rate', 0):.1f}%"],
            ["Depreciation Rate", f"{assumptions.get('depreciation_rate', 0):.1f}%"],
            ["CapEx Rate", f"{assumptions.get('capex_rate', 0):.1f}%"],
            ["Working Capital Change", f"{assumptions.get('working_capital_change', 0):.1f}%"],
            ["Risk-Free Rate", f"{assumptions.get('risk_free_rate', 0):.1f}%"],
            ["Market Risk Premium", f"{assumptions.get('market_risk_premium', 0):.1f}%"],
            ["Beta", f"{assumptions.get('beta', 0):.2f}"],
            ["Debt Ratio", f"{assumptions.get('debt_ratio', 0):.1f}%"],
            ["Cost of Debt", f"{assumptions.get('cost_of_debt', 0):.1f}%"],
            ["Terminal Growth Rate", f"{assumptions.get('terminal_growth_rate', 0):.1f}%"],
        ]
        
        for assumption, value in financial_data:
            ws[f'A{row}'] = assumption
            ws[f'B{row}'] = value
            row += 1
        
        # Apply borders
        for row_cells in ws.iter_rows(min_row=3, max_row=row-1, min_col=1, max_col=2):
            for cell in row_cells:
                cell.border = self.border
    
    def _create_projections_sheet(self, ws, projections: Dict):
        """Create financial projections worksheet"""
        
        ws['A1'] = "FINANCIAL PROJECTIONS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:G1')
        
        # Headers
        headers = ['Year'] + [f'Year {i}' for i in projections.get('years', range(1, 6))]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Financial data rows
        financial_rows = [
            ('Revenue ($M)', projections.get('revenue', [])),
            ('Gross Profit ($M)', projections.get('gross_profit', [])),
            ('Operating Profit ($M)', projections.get('operating_profit', [])),
            ('EBIT ($M)', projections.get('ebit', [])),
            ('Taxes ($M)', projections.get('taxes', [])),
            ('NOPAT ($M)', projections.get('nopat', [])),
            ('Depreciation ($M)', projections.get('depreciation', [])),
            ('CapEx ($M)', projections.get('capex', [])),
            ('Working Capital Change ($M)', projections.get('working_capital_change', [])),
            ('Free Cash Flow ($M)', projections.get('free_cash_flow', [])),
        ]
        
        for row_idx, (label, values) in enumerate(financial_rows, start=4):
            ws.cell(row=row_idx, column=1).value = label
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            
            for col_idx, value in enumerate(values, start=2):
                if isinstance(value, (int, float)):
                    # Convert to millions if the value is large
                    display_value = value / 1_000_000 if abs(value) > 1_000_000 else value
                    ws.cell(row=row_idx, column=col_idx).value = display_value
                    ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.0'
                else:
                    ws.cell(row=row_idx, column=col_idx).value = value
        
        # Apply borders to all data
        max_row = 4 + len(financial_rows) - 1
        max_col = len(headers)
        for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=max_col):
            for cell in row:
                cell.border = self.border
                if cell.column > 1:  # Numeric columns
                    cell.alignment = Alignment(horizontal='right')
        
        # Add chart
        self._add_projections_chart(ws, len(headers), len(financial_rows))
    
    def _create_dcf_sheet(self, ws, dcf_model: Dict):
        """Create DCF calculation worksheet"""
        
        ws['A1'] = "DCF VALUATION CALCULATION"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        # WACC calculation
        ws['A3'] = "WACC CALCULATION"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill
        ws.merge_cells('A3:C3')
        
        wacc_data = [
            ["Cost of Equity", f"{dcf_model['cost_of_equity']:.2f}%"],
            ["Cost of Debt", f"{dcf_model['cost_of_debt']:.2f}%"],
            ["Tax Rate", f"{dcf_model['tax_rate']:.2f}%"],
            ["Debt-to-Equity Ratio", f"{dcf_model['debt_to_equity_ratio']:.2f}"],
            ["WACC", f"{dcf_model['wacc']:.2f}%"],
        ]
        
        row = 4
        for item, value in wacc_data:
            ws[f'A{row}'] = item
            ws[f'B{row}'] = value
            if item == "WACC":
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        row += 1
        
        # Terminal value calculation
        ws[f'A{row}'] = "TERMINAL VALUE CALCULATION"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        terminal_data = [
            ["Terminal Growth Rate", f"{dcf_model['terminal_growth_rate']:.2f}%"],
            ["Terminal Value", f"${dcf_model['terminal_value']:,.0f}"],
        ]
        
        for item, value in terminal_data:
            ws[f'A{row}'] = item
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Valuation summary
        ws[f'A{row}'] = "VALUATION SUMMARY"
        ws[f'A{row}'].font = self.header_font
        ws[f'A{row}'].fill = self.header_fill
        ws.merge_cells(f'A{row}:C{row}')
        row += 1
        
        valuation_data = [
            ["PV of Free Cash Flows", f"${dcf_model['pv_of_fcf']:,.0f}"],
            ["PV of Terminal Value", f"${dcf_model['pv_of_terminal_value']:,.0f}"],
            ["Enterprise Value", f"${dcf_model['enterprise_value']:,.0f}"],
            ["Equity Value", f"${dcf_model['equity_value']:,.0f}"],
        ]
        
        if dcf_model.get('shares_outstanding') and dcf_model.get('value_per_share'):
            valuation_data.extend([
                ["Shares Outstanding", f"{dcf_model['shares_outstanding']:,.0f}"],
                ["Value per Share", f"${dcf_model['value_per_share']:.2f}"],
            ])
        
        for item, value in valuation_data:
            ws[f'A{row}'] = item
            ws[f'B{row}'] = value
            if item in ["Enterprise Value", "Equity Value", "Value per Share"]:
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
            row += 1
        
        # Apply borders
        for row_cells in ws.iter_rows(min_row=3, max_row=row-1, min_col=1, max_col=2):
            for cell in row_cells:
                cell.border = self.border
    
    def _create_sensitivity_sheet(self, ws, financial_model: Dict, dcf_model: Dict):
        """Create sensitivity analysis worksheet"""
        
        ws['A1'] = "SENSITIVITY ANALYSIS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Create sample sensitivity table for WACC vs Terminal Growth Rate
        ws['A3'] = "Equity Value Sensitivity (WACC vs Terminal Growth Rate)"
        ws['A3'].font = Font(bold=True)
        ws.merge_cells('A3:F3')
        
        # Headers
        ws['B5'] = "Terminal Growth Rate"
        ws.merge_cells('B5:F5')
        ws['B5'].font = self.header_font
        ws['B5'].fill = self.header_fill
        ws['B5'].alignment = self.center_alignment
        
        # Terminal growth rates
        terminal_rates = [1.5, 2.0, 2.5, 3.0, 3.5]
        for i, rate in enumerate(terminal_rates, start=2):
            cell = ws.cell(row=6, column=i)
            cell.value = f"{rate}%"
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_alignment
            cell.border = self.border
        
        # WACC label
        ws['A6'] = "WACC"
        ws['A6'].font = self.header_font
        ws['A6'].fill = self.header_fill
        ws['A6'].alignment = self.center_alignment
        ws['A6'].border = self.border
        
        # WACC rates and sensitivity values
        base_wacc = dcf_model['wacc']
        wacc_rates = [base_wacc - 1, base_wacc - 0.5, base_wacc, base_wacc + 0.5, base_wacc + 1]
        base_equity_value = dcf_model['equity_value']
        
        for i, wacc in enumerate(wacc_rates, start=7):
            # WACC label
            ws.cell(row=i, column=1).value = f"{wacc:.1f}%"
            ws.cell(row=i, column=1).font = Font(bold=True)
            ws.cell(row=i, column=1).border = self.border
            ws.cell(row=i, column=1).alignment = self.center_alignment
            
            # Sensitivity values (simplified calculation)
            for j, terminal_rate in enumerate(terminal_rates, start=2):
                # Simple sensitivity approximation
                wacc_impact = (base_wacc - wacc) * 0.15  # 15% impact per 1% WACC change
                terminal_impact = (terminal_rate - dcf_model['terminal_growth_rate']) * 0.10  # 10% impact per 1% terminal change
                
                adjusted_value = base_equity_value * (1 + (wacc_impact + terminal_impact))
                
                cell = ws.cell(row=i, column=j)
                cell.value = adjusted_value
                cell.number_format = '#,##0'
                cell.border = self.border
                cell.alignment = Alignment(horizontal='right')
                
                # Color coding
                if adjusted_value > base_equity_value * 1.1:
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif adjusted_value < base_equity_value * 0.9:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    def _create_scenarios_sheet(self, ws, scenarios: List[Dict]):
        """Create scenario analysis worksheet"""
        
        ws['A1'] = "SCENARIO ANALYSIS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        # Headers
        headers = ["Scenario", "Type", "Probability", "Key Changes", "Equity Value", "vs Base Case"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Scenario data
        for row_idx, scenario in enumerate(scenarios, start=4):
            ws.cell(row=row_idx, column=1).value = scenario['scenario_name']
            ws.cell(row=row_idx, column=2).value = scenario['scenario_type'].replace('_', ' ').title()
            ws.cell(row=row_idx, column=3).value = f"{scenario.get('probability', 0)}%"
            
            # Key changes summary
            changes = scenario.get('assumption_changes', {})
            change_summary = ', '.join([f"{k}: {v}" for k, v in list(changes.items())[:2]])
            ws.cell(row=row_idx, column=4).value = change_summary
            
            # Results (simplified)
            equity_value = scenario.get('scenario_results', {}).get('equity_value', 0)
            ws.cell(row=row_idx, column=5).value = equity_value
            ws.cell(row=row_idx, column=5).number_format = '#,##0'
            
            # Variance (simplified)
            variance = scenario.get('variance_from_base', {}).get('equity_value_change', 0)
            ws.cell(row=row_idx, column=6).value = f"{variance:.1f}%"
        
        # Apply borders
        max_row = 4 + len(scenarios) - 1
        for row in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=6):
            for cell in row:
                cell.border = self.border
    
    def _add_projections_chart(self, ws, num_cols: int, num_rows: int):
        """Add chart to projections worksheet"""
        
        # Create chart
        chart = LineChart()
        chart.title = "Financial Projections"
        chart.style = 10
        chart.x_axis.title = "Years"
        chart.y_axis.title = "Value ($M)"
        
        # Add data series for Revenue and FCF
        revenue_ref = Reference(ws, min_col=2, min_row=4, max_col=num_cols, max_row=4)
        fcf_ref = Reference(ws, min_col=2, min_row=4+9, max_col=num_cols, max_row=4+9)  # FCF is 10th row
        categories = Reference(ws, min_col=2, min_row=3, max_col=num_cols, max_row=3)
        
        chart.add_data(revenue_ref, titles_from_data=False)
        chart.add_data(fcf_ref, titles_from_data=False)
        chart.set_categories(categories)
        
        # Series names
        chart.series[0].title = "Revenue"
        chart.series[1].title = "Free Cash Flow"
        
        # Add chart to worksheet
        ws.add_chart(chart, f"A{4 + num_rows + 2}")
    
    def export_forecast_to_excel(self, forecast: Dict[str, Any], predictions: List[Dict[str, Any]]) -> BytesIO:
        """Export forecast model to Excel"""
        
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Create worksheets
        summary_ws = wb.create_sheet("Forecast Summary")
        data_ws = wb.create_sheet("Historical & Forecast Data")
        accuracy_ws = wb.create_sheet("Model Performance")
        
        # Populate summary
        self._create_forecast_summary(summary_ws, forecast)
        
        # Populate data
        self._create_forecast_data_sheet(data_ws, forecast, predictions)
        
        # Populate accuracy
        self._create_accuracy_sheet(accuracy_ws, forecast)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    
    def _create_forecast_summary(self, ws, forecast: Dict):
        """Create forecast summary worksheet"""
        
        ws['A1'] = f"Forecast: {forecast['forecast_name']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Forecast details
        details = [
            ["Forecast Type", forecast['forecast_type'].title()],
            ["Method", forecast['forecast_method'].replace('_', ' ').title()],
            ["Frequency", forecast['frequency'].title()],
            ["Base Period", f"{forecast['base_period_start']} to {forecast['base_period_end']}"],
            ["Forecast Period", f"{forecast['forecast_start']} to {forecast['forecast_end']}"],
            ["Status", forecast['status'].title()],
            ["Created", forecast['created_at'][:10]],
        ]
        
        row = 3
        for label, value in details:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Accuracy metrics if available
        if forecast.get('accuracy_metrics'):
            row += 1
            ws[f'A{row}'] = "MODEL PERFORMANCE"
            ws[f'A{row}'].font = self.header_font
            ws[f'A{row}'].fill = self.header_fill
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
            
            metrics = forecast['accuracy_metrics']
            accuracy_data = [
                ["MAE", f"{metrics.get('mae', 0):,.2f}"],
                ["RMSE", f"{metrics.get('rmse', 0):,.2f}"],
                ["MAPE", f"{metrics.get('mape', 0):.2f}%"],
            ]
            
            for label, value in accuracy_data:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = value
                row += 1
    
    def _create_forecast_data_sheet(self, ws, forecast: Dict, predictions: List[Dict]):
        """Create forecast data worksheet"""
        
        ws['A1'] = "HISTORICAL & FORECAST DATA"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ["Date", "Historical Value", "Predicted Value", "Lower Bound", "Upper Bound"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col)
            cell.value = header
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment
        
        # Historical data
        historical_data = forecast.get('historical_data', {})
        dates = historical_data.get('dates', [])
        values = historical_data.get('values', [])
        
        row = 4
        for date_str, value in zip(dates, values):
            ws.cell(row=row, column=1).value = date_str
            ws.cell(row=row, column=2).value = value
            ws.cell(row=row, column=2).number_format = '#,##0.0'
            row += 1
        
        # Forecast data
        forecast_data = forecast.get('forecast_data', {})
        forecast_dates = forecast_data.get('dates', [])
        forecast_values = forecast_data.get('values', [])
        
        # Confidence intervals if available
        confidence = forecast.get('confidence_intervals', {})
        lower_bounds = confidence.get('lower_bound', [])
        upper_bounds = confidence.get('upper_bound', [])
        
        for i, (date_str, value) in enumerate(zip(forecast_dates, forecast_values)):
            ws.cell(row=row, column=1).value = date_str
            ws.cell(row=row, column=3).value = value
            ws.cell(row=row, column=3).number_format = '#,##0.0'
            
            if i < len(lower_bounds):
                ws.cell(row=row, column=4).value = lower_bounds[i]
                ws.cell(row=row, column=4).number_format = '#,##0.0'
            
            if i < len(upper_bounds):
                ws.cell(row=row, column=5).value = upper_bounds[i]
                ws.cell(row=row, column=5).number_format = '#,##0.0'
            
            row += 1
        
        # Apply borders
        max_row = row - 1
        for row_cells in ws.iter_rows(min_row=3, max_row=max_row, min_col=1, max_col=5):
            for cell in row_cells:
                cell.border = self.border
    
    def _create_accuracy_sheet(self, ws, forecast: Dict):
        """Create model accuracy worksheet"""
        
        ws['A1'] = "MODEL PERFORMANCE ANALYSIS"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        accuracy_metrics = forecast.get('accuracy_metrics', {})
        
        if not accuracy_metrics:
            ws['A3'] = "No accuracy metrics available yet."
            return
        
        # Performance summary
        ws['A3'] = "ACCURACY METRICS"
        ws['A3'].font = self.header_font
        ws['A3'].fill = self.header_fill
        ws.merge_cells('A3:C3')
        
        metrics_data = [
            ["Mean Absolute Error (MAE)", f"{accuracy_metrics.get('mae', 0):,.2f}"],
            ["Root Mean Square Error (RMSE)", f"{accuracy_metrics.get('rmse', 0):,.2f}"],
            ["Mean Absolute Percentage Error (MAPE)", f"{accuracy_metrics.get('mape', 0):.2f}%"],
            ["Directional Accuracy", f"{accuracy_metrics.get('directional_accuracy', 0):.1f}%"],
        ]
        
        row = 4
        for metric, value in metrics_data:
            ws[f'A{row}'] = metric
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Feature importance if available
        feature_importance = accuracy_metrics.get('feature_importance', {})
        if feature_importance:
            row += 1
            ws[f'A{row}'] = "FEATURE IMPORTANCE"
            ws[f'A{row}'].font = self.header_font
            ws[f'A{row}'].fill = self.header_fill
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
            
            for feature, importance in feature_importance.items():
                ws[f'A{row}'] = feature.replace('_', ' ').title()
                ws[f'B{row}'] = f"{importance:.3f}"
                row += 1
    
    def export_to_csv(self, data: List[Dict[str, Any]], filename: str = "export") -> BytesIO:
        """Export data to CSV format"""
        
        if not data:
            raise ValueError("No data provided for CSV export")
        
        # Convert to pandas DataFrame
        df = pd.DataFrame(data)
        
        # Convert datetime and decimal objects to strings
        for col in df.columns:
            if df[col].dtype == 'object':
                df[col] = df[col].apply(lambda x: str(x) if isinstance(x, (datetime, date, Decimal)) else x)
        
        # Export to CSV
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8')
        output.seek(0)
        
        return output