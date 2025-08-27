# app/services/excel_service.py

import io
import logging
import pandas as pd
from typing import List, Dict, Optional
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

logger = logging.getLogger(__name__)

class ExcelService:
    @staticmethod
    async def validate_excel_file(file, required_columns: List[str], sheet_name: str = None) -> Dict:
        """
        Validate Excel file and return validation results without processing data
        
        Args:
            file: Uploaded file object
            required_columns: List of required column names
            sheet_name: Optional sheet name
            
        Returns:
            Dictionary with validation results and file info
        """
        try:
            # Read the uploaded file content
            content = await file.read()
            excel_buffer = io.BytesIO(content)
            
            # Basic file info
            file_info = {
                "filename": file.filename,
                "size_bytes": len(content),
                "file_type": "xlsx" if file.filename.endswith('.xlsx') else "xls"
            }
            
            # Try to determine the appropriate data sheet name if not provided
            sheet_names = []
            try:
                xl_file = pd.ExcelFile(excel_buffer)
                sheet_names = xl_file.sheet_names
                file_info["sheet_names"] = sheet_names
                
                if sheet_name is None:
                    # Look for common data sheet patterns
                    data_sheet_patterns = [
                        "Stock Import Template", "Product Import Template", "Vendor Import Template", 
                        "Customer Import Template", "Import Template", "Data", "Sheet1"
                    ]
                    
                    for pattern in data_sheet_patterns:
                        if pattern in sheet_names:
                            sheet_name = pattern
                            break
                    
                    # If no pattern match, use the first sheet
                    if sheet_name is None and sheet_names:
                        sheet_name = sheet_names[0]
                        
            except Exception as e:
                return {
                    "valid": False,
                    "file_info": file_info,
                    "validation_errors": [f"Cannot read Excel file structure: {str(e)}"],
                    "validation_warnings": [],
                    "preview_data": [],
                    "total_rows": 0
                }
            
            # Reset buffer position
            excel_buffer.seek(0)
            
            try:
                # Read the data
                df = pd.read_excel(excel_buffer, sheet_name=sheet_name, 
                                   engine='openpyxl' if file.filename.endswith('.xlsx') else 'xlrd')
                
                # Clean column names
                original_columns = df.columns.tolist()
                df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
                
                file_info["columns_found"] = original_columns
                file_info["total_rows"] = len(df)
                file_info["data_sheet"] = sheet_name
                
                validation_errors = []
                validation_warnings = []
                
                # Check for empty file
                if df.empty:
                    validation_errors.append("Excel file contains no data rows")
                
                # Validate required columns
                normalized_required = [col.lower().replace(' ', '_') for col in required_columns]
                missing_columns = [col for col in required_columns 
                                   if col.lower().replace(' ', '_') not in df.columns]
                
                if missing_columns:
                    validation_errors.append(
                        f"Missing required columns: {', '.join(missing_columns)}. "
                        f"Found columns: {', '.join(original_columns)}"
                    )
                
                # Check for completely empty rows
                empty_rows = df.dropna(how='all')
                if len(empty_rows) < len(df):
                    validation_warnings.append(
                        f"{len(df) - len(empty_rows)} empty rows will be skipped during import"
                    )
                
                # Generate preview data (first 5 rows)
                preview_data = []
                if not df.empty:
                    preview_df = df.head(5)
                    for idx, row in preview_df.iterrows():
                        row_dict = row.to_dict()
                        # Convert nan values to None for JSON serialization
                        row_dict = {k: (None if pd.isna(v) else v) for k, v in row_dict.items()}
                        row_dict["_row_number"] = idx + 2  # +2 because Excel is 1-indexed and has header
                        preview_data.append(row_dict)
                
                return {
                    "valid": len(validation_errors) == 0,
                    "file_info": file_info,
                    "validation_errors": validation_errors,
                    "validation_warnings": validation_warnings,
                    "preview_data": preview_data,
                    "total_rows": len(df)
                }
                
            except Exception as e:
                return {
                    "valid": False,
                    "file_info": file_info,
                    "validation_errors": [f"Error reading Excel data: {str(e)}"],
                    "validation_warnings": [],
                    "preview_data": [],
                    "total_rows": 0
                }
                
        except Exception as e:
            logger.error(f"Excel validation error: {str(e)}")
            return {
                "valid": False,
                "file_info": {"filename": file.filename if hasattr(file, 'filename') else 'unknown'},
                "validation_errors": [f"Failed to process Excel file: {str(e)}"],
                "validation_warnings": [],
                "preview_data": [],
                "total_rows": 0
            }
    @staticmethod
    async def parse_excel_file(file, required_columns: List[str], sheet_name: str = None) -> List[Dict]:
        """
        Parse Excel file and return list of dictionaries.
        Supports both .xlsx and .xls formats.
        """
        try:
            # Read the uploaded file content
            content = await file.read()
            excel_buffer = io.BytesIO(content)
            
            # Try to determine the appropriate data sheet name if not provided
            if sheet_name is None:
                # Try to read all sheet names and find the data sheet
                try:
                    xl_file = pd.ExcelFile(excel_buffer)
                    sheet_names = xl_file.sheet_names
                    
                    # Look for common data sheet patterns
                    data_sheet_patterns = [
                        "Stock Import Template", "Product Import Template", "Vendor Import Template", 
                        "Customer Import Template", "Import Template", "Data", "Sheet1"
                    ]
                    
                    sheet_name = None
                    for pattern in data_sheet_patterns:
                        if pattern in sheet_names:
                            sheet_name = pattern
                            break
                    
                    # If no pattern match, use the first sheet
                    if sheet_name is None and sheet_names:
                        sheet_name = sheet_names[0]
                        
                except Exception:
                    # If we can't read sheet names, try the default approach
                    sheet_name = 0  # First sheet
            
            # Reset buffer position
            excel_buffer.seek(0)
            
            # Use pandas to read Excel, specifying the sheet name for data
            df = pd.read_excel(excel_buffer, sheet_name=sheet_name, 
                               engine='openpyxl' if file.filename.endswith('.xlsx') else 'xlrd')
            
            # Clean column names
            df.columns = df.columns.str.strip().str.lower().str.replace(' ', '_')
            
            # Validate required columns
            missing_columns = [col for col in required_columns if col.lower().replace(' ', '_') not in df.columns]
            if missing_columns:
                found_columns = ', '.join(df.columns)
                sheet_info = f"sheet '{sheet_name}'" if isinstance(sheet_name, str) else f"sheet {sheet_name}"
                raise ValueError(f"Missing required columns in {sheet_info}: {', '.join(missing_columns)}. "
                               f"Found columns: {found_columns}. "
                               f"Make sure to upload a data file with the correct sheet and headers, not the instructions sheet.")
            
            # Convert to list of dicts, handling NaN values
            records = df.replace({pd.NA: None, float('nan'): None}).to_dict(orient='records')
            
            logger.info(f"Successfully parsed {len(records)} records from Excel file")
            return records
            
        except ValueError as ve:
            logger.error(f"Excel validation error: {str(ve)}")
            raise
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}")
            raise ValueError(f"Invalid Excel file format or error reading data sheet: {str(e)}. "
                           f"Please use the downloaded template and fill in the data sheet.")

    @staticmethod
    def create_streaming_response(excel_data: io.BytesIO, filename: str) -> StreamingResponse:
        """Create streaming response for Excel download"""
        headers = {
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        return StreamingResponse(
            iter([excel_data.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )

class StockExcelService(ExcelService):
    REQUIRED_COLUMNS = [
        "Product Name",
        "Quantity",
        "Unit",
        "HSN Code",
        "Part Number",
        "Unit Price",
        "GST Rate",
        "Reorder Level",
        "Location"
    ]

    @staticmethod
    def create_template() -> io.BytesIO:
        """Create Excel template for stock import with styling and sample data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Import Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = StockExcelService.REQUIRED_COLUMNS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add sample data
        sample_data = [
            "Steel Bolt M8x50",  # Product Name
            100,                 # Quantity
            "PCS",               # Unit
            "73181590",          # HSN Code
            "SB-M8-50",          # Part Number
            25.50,               # Unit Price
            18.0,                # GST Rate
            50,                  # Reorder Level
            "Warehouse A-1"      # Location
        ]

        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        instructions = [
            "Instructions for Stock Import:",
            "1. Required columns must be present and spelled exactly as in the template.",
            "2. Product Name is mandatory and must be unique per organization.",
            "3. If product doesn't exist, it will be created automatically.",
            "4. Quantity must be a non-negative number.",
            "5. Unit Price and GST Rate should be numbers.",
            "6. Reorder Level should be an integer.",
            "7. Location is optional.",
            "8. Do not modify the header row.",
            "9. Save as .xlsx format."
        ]

        for row, text in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=text)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("Stock Excel template generated successfully")
        return excel_buffer

    @staticmethod
    def export_stock(stock_data: List[Dict]) -> io.BytesIO:
        """Export stock data to Excel with styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Stock Export"

        # Define styles (same as template)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = [
            "Product Name", "Quantity", "Unit", "HSN Code", "Part Number",
            "Unit Price", "GST Rate", "Reorder Level", "Location"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add data
        for row_idx, item in enumerate(stock_data, 2):
            row_data = [
                item.get("product_name"),
                item.get("quantity"),
                item.get("unit"),
                item.get("hsn_code"),
                item.get("part_number"),
                item.get("unit_price"),
                item.get("gst_rate"),
                item.get("reorder_level"),
                item.get("location")
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Stock data exported successfully: {len(stock_data)} records")
        return excel_buffer

class VendorExcelService(ExcelService):
    REQUIRED_COLUMNS = [
        "Name",
        "Contact Number",
        "Email",
        "Address Line 1",
        "Address Line 2",
        "City",
        "State",
        "Pin Code",
        "State Code",
        "GST Number",
        "PAN Number"
    ]

    @staticmethod
    def create_template() -> io.BytesIO:
        """Create Excel template for vendor import with styling and sample data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendor Import Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = VendorExcelService.REQUIRED_COLUMNS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add sample data
        sample_data = [
            "Test Vendor Inc",     # Name
            "1234567890",          # Contact Number
            "test@vendor.com",     # Email
            "123 Vendor Street",   # Address Line 1
            "Suite 456",           # Address Line 2
            "Vendor City",         # City
            "Vendor State",        # State
            "123456",              # Pin Code
            "27",                  # State Code
            "27AACFV1234D1Z5",     # GST Number
            "AACFV1234D"           # PAN Number
        ]

        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        instructions = [
            "Instructions for Vendor Import:",
            "1. Required columns must be present and spelled exactly as in the template.",
            "2. Name and Contact Number are mandatory.",
            "3. Email, Address Line 2, GST Number, and PAN Number are optional.",
            "4. Pin Code and State Code must be valid.",
            "5. If vendor name exists, it will be updated; otherwise, created.",
            "6. Do not modify the header row.",
            "7. Save as .xlsx format."
        ]

        for row, text in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=text)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("Vendor Excel template generated successfully")
        return excel_buffer

    @staticmethod
    def export_vendors(vendors_data: List[Dict]) -> io.BytesIO:
        """Export vendors data to Excel with styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Vendors Export"

        # Define styles (same as template)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = [
            "Name", "Contact Number", "Email", "Address Line 1", "Address Line 2",
            "City", "State", "Pin Code", "State Code", "GST Number", "PAN Number"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add data
        for row_idx, item in enumerate(vendors_data, 2):
            row_data = [
                item.get("name"),
                item.get("contact_number"),
                item.get("email"),
                item.get("address1"),
                item.get("address2"),
                item.get("city"),
                item.get("state"),
                item.get("pin_code"),
                item.get("state_code"),
                item.get("gst_number"),
                item.get("pan_number")
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Vendors data exported successfully: {len(vendors_data)} records")
        return excel_buffer

class CustomerExcelService(ExcelService):
    REQUIRED_COLUMNS = [
        "Name",
        "Contact Number",
        "Email",
        "Address Line 1",
        "Address Line 2",
        "City",
        "State",
        "Pin Code",
        "State Code",
        "GST Number",
        "PAN Number"
    ]

    @staticmethod
    def create_template() -> io.BytesIO:
        """Create Excel template for customer import with styling and sample data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Customer Import Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = CustomerExcelService.REQUIRED_COLUMNS
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add sample data
        sample_data = [
            "Test Customer LLC",   # Name
            "9876543210",          # Contact Number
            "test@customer.com",   # Email
            "456 Customer Ave",    # Address Line 1
            "Apt 789",             # Address Line 2
            "Customer City",       # City
            "Customer State",      # State
            "654321",              # Pin Code
            "29",                  # State Code
            "29AAACC1234E1Z7",     # GST Number
            "AAACC1234E"           # PAN Number
        ]

        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        instructions = [
            "Instructions for Customer Import:",
            "1. Required columns must be present and spelled exactly as in the template.",
            "2. Name and Contact Number are mandatory.",
            "3. Email, Address Line 2, GST Number, and PAN Number are optional.",
            "4. Pin Code and State Code must be valid.",
            "5. If customer name exists, it will be updated; otherwise, created.",
            "6. Do not modify the header row.",
            "7. Save as .xlsx format."
        ]

        for row, text in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=text)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("Customer Excel template generated successfully")
        return excel_buffer

    @staticmethod
    def export_customers(customers_data: List[Dict]) -> io.BytesIO:
        """Export customers data to Excel with styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Customers Export"

        # Define styles (same as template)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = [
            "Name", "Contact Number", "Email", "Address Line 1", "Address Line 2",
            "City", "State", "Pin Code", "State Code", "GST Number", "PAN Number"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add data
        for row_idx, item in enumerate(customers_data, 2):
            row_data = [
                item.get("name"),
                item.get("contact_number"),
                item.get("email"),
                item.get("address1"),
                item.get("address2"),
                item.get("city"),
                item.get("state"),
                item.get("pin_code"),
                item.get("state_code"),
                item.get("gst_number"),
                item.get("pan_number")
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Customers data exported successfully: {len(customers_data)} records")
        return excel_buffer

class ProductExcelService(ExcelService):
    REQUIRED_COLUMNS = [
        "Product Name",
        "HSN Code",
        "Part Number",
        "Unit",
        "Unit Price",
        "GST Rate",
        "Is GST Inclusive",
        "Reorder Level",
        "Description",
        "Is Manufactured"
    ]

    @staticmethod
    def create_template() -> io.BytesIO:
        """Create Excel template for product import with styling and sample data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Product Import Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers (including optional initial stock columns)
        headers = ProductExcelService.REQUIRED_COLUMNS + ["Initial Quantity", "Initial Location"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add sample data
        sample_data = [
            "Test Product",        # Product Name
            "123456",              # HSN Code
            "TP-001",              # Part Number
            "PCS",                 # Unit
            100.0,                 # Unit Price
            18.0,                  # GST Rate
            "FALSE",               # Is GST Inclusive
            10,                    # Reorder Level
            "Test description",    # Description
            "FALSE",               # Is Manufactured
            50,                    # Initial Quantity (optional)
            "Warehouse A"          # Initial Location (optional)
        ]

        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        instructions = [
            "Instructions for Product Import:",
            "1. Required columns must be present and spelled exactly as in the template.",
            "2. Product Name and Unit are mandatory.",
            "3. HSN Code, Part Number, Description are optional but recommended.",
            "4. Unit Price, GST Rate, Reorder Level should be numbers.",
            "5. Is GST Inclusive and Is Manufactured should be TRUE/FALSE or YES/NO.",
            "6. Initial Quantity and Initial Location are optional for initial stock setup.",
            "7. If product name exists, it will be updated; otherwise, created.",
            "8. Do not modify the header row.",
            "9. Save as .xlsx format."
        ]

        for row, text in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=text)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("Product Excel template generated successfully")
        return excel_buffer

    @staticmethod
    def export_products(products_data: List[Dict]) -> io.BytesIO:
        """Export products data to Excel with styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Products Export"

        # Define styles (same as template)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = [
            "Product Name", "HSN Code", "Part Number", "Unit",
            "Unit Price", "GST Rate", "Is GST Inclusive", "Reorder Level",
            "Description", "Is Manufactured"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add data
        for row_idx, item in enumerate(products_data, 2):
            row_data = [
                item.get("product_name"),
                item.get("hsn_code"),
                item.get("part_number"),
                item.get("unit"),
                item.get("unit_price"),
                item.get("gst_rate"),
                item.get("is_gst_inclusive"),
                item.get("reorder_level"),
                item.get("description"),
                item.get("is_manufactured")
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Products data exported successfully: {len(products_data)} records")
        return excel_buffer

class ReportsExcelService(ExcelService):
    """Service for exporting various reports to Excel"""
    
    @staticmethod
    def export_sales_report(sales_data: List[Dict]) -> io.BytesIO:
        """Export sales report to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sales Report"
        
        # Headers
        headers = [
            "Voucher Number", "Date", "Customer Name", "Total Amount", 
            "GST Amount", "Net Amount", "Status"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for voucher in sales_data.get('vouchers', []):
            ws.append([
                voucher.get('voucher_number', ''),
                voucher.get('date', ''),
                voucher.get('customer_name', ''),
                voucher.get('total_amount', 0),
                voucher.get('gst_amount', 0),
                voucher.get('total_amount', 0) - voucher.get('gst_amount', 0),
                voucher.get('status', '')
            ])
        
        # Add summary
        if 'summary' in sales_data:
            summary = sales_data['summary']
            ws.append([])  # Empty row
            ws.append(['SUMMARY'])
            ws.append(['Total Vouchers', summary.get('total_vouchers', 0)])
            ws.append(['Total Sales', summary.get('total_sales', 0)])
            ws.append(['Total GST', summary.get('total_gst', 0)])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_data = io.BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        return excel_data
    
    @staticmethod
    def export_purchase_report(purchase_data: List[Dict]) -> io.BytesIO:
        """Export purchase report to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Purchase Report"
        
        # Headers
        headers = [
            "Voucher Number", "Date", "Vendor Name", "Total Amount", 
            "GST Amount", "Net Amount", "Status"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for voucher in purchase_data.get('vouchers', []):
            ws.append([
                voucher.get('voucher_number', ''),
                voucher.get('date', ''),
                voucher.get('vendor_name', ''),
                voucher.get('total_amount', 0),
                voucher.get('gst_amount', 0),
                voucher.get('total_amount', 0) - voucher.get('gst_amount', 0),
                voucher.get('status', '')
            ])
        
        # Add summary
        if 'summary' in purchase_data:
            summary = purchase_data['summary']
            ws.append([])  # Empty row
            ws.append(['SUMMARY'])
            ws.append(['Total Vouchers', summary.get('total_vouchers', 0)])
            ws.append(['Total Purchases', summary.get('total_purchases', 0)])
            ws.append(['Total GST', summary.get('total_gst', 0)])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_data = io.BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        return excel_data
    
    @staticmethod
    def export_inventory_report(inventory_data: List[Dict]) -> io.BytesIO:
        """Export inventory report to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory Report"
        
        # Headers
        headers = [
            "Product Name", "HSN Code", "Current Stock", "Unit", 
            "Unit Price", "Total Value", "Reorder Level", "Status"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for item in inventory_data.get('items', []):
            status = "Low Stock" if item.get('quantity', 0) <= item.get('reorder_level', 0) else "In Stock"
            ws.append([
                item.get('product_name', ''),
                item.get('hsn_code', ''),
                item.get('quantity', 0),
                item.get('unit', ''),
                item.get('unit_price', 0),
                item.get('quantity', 0) * item.get('unit_price', 0),
                item.get('reorder_level', 0),
                status
            ])
        
        # Add summary
        if 'summary' in inventory_data:
            summary = inventory_data['summary']
            ws.append([])  # Empty row
            ws.append(['SUMMARY'])
            ws.append(['Total Products', summary.get('total_products', 0)])
            ws.append(['Total Stock Value', summary.get('total_value', 0)])
            ws.append(['Low Stock Items', summary.get('low_stock_items', 0)])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_data = io.BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        return excel_data
    
    @staticmethod
    def export_pending_orders_report(orders_data: List[Dict]) -> io.BytesIO:
        """Export pending orders report to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Pending Orders"
        
        # Headers
        headers = [
            "Order Number", "Type", "Date", "Party Name", "Total Amount", 
            "Status", "Days Pending"
        ]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for order in orders_data.get('orders', []):
            ws.append([
                order.get('order_number', ''),
                order.get('order_type', ''),
                order.get('date', ''),
                order.get('party_name', ''),
                order.get('total_amount', 0),
                order.get('status', ''),
                order.get('days_pending', 0)
            ])
        
        # Add summary
        if 'summary' in orders_data:
            summary = orders_data['summary']
            ws.append([])  # Empty row
            ws.append(['SUMMARY'])
            ws.append(['Total Orders', summary.get('total_orders', 0)])
            ws.append(['Total Value', summary.get('total_value', 0)])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_data = io.BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        return excel_data
    
    @staticmethod
    def export_ledger_report(ledger_data: Dict, report_type: str = "complete") -> io.BytesIO:
        """Export ledger report to Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = f"{report_type.title()} Ledger"
        
        if report_type == "complete":
            # Complete Ledger Headers
            headers = [
                "Date", "Voucher Type", "Voucher Number", "Account Type", 
                "Account Name", "Debit Amount", "Credit Amount", "Running Balance", "Description"
            ]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for transaction in ledger_data.get('transactions', []):
                ws.append([
                    transaction.get('date', ''),
                    transaction.get('voucher_type', ''),
                    transaction.get('voucher_number', ''),
                    transaction.get('account_type', ''),
                    transaction.get('account_name', ''),
                    transaction.get('debit_amount', 0),
                    transaction.get('credit_amount', 0),
                    transaction.get('running_balance', 0),
                    transaction.get('description', '')
                ])
        else:
            # Outstanding Ledger Headers
            headers = [
                "Account Type", "Account Name", "Outstanding Amount", 
                "Last Transaction", "Transaction Count", "Contact"
            ]
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Add data
            for balance in ledger_data.get('outstanding_balances', []):
                ws.append([
                    balance.get('account_type', ''),
                    balance.get('account_name', ''),
                    balance.get('outstanding_amount', 0),
                    balance.get('last_transaction_date', ''),
                    balance.get('transaction_count', 0),
                    balance.get('contact_info', '')
                ])
        
        # Add summary
        if 'summary' in ledger_data:
            summary = ledger_data['summary']
            ws.append([])  # Empty row
            ws.append(['SUMMARY'])
            for key, value in summary.items():
                ws.append([key.replace('_', ' ').title(), value])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        excel_data = io.BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        return excel_data


class CompanyExcelService(ExcelService):
    REQUIRED_COLUMNS = [
        "Name",
        "Address Line 1", 
        "City",
        "State",
        "Pin Code",
        "State Code",
        "Contact Number"
    ]

    @staticmethod
    def create_template() -> io.BytesIO:
        """Create Excel template for company import with styling and sample data"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Company Import Template"

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers (including optional fields)
        headers = [
            "Name", "Address Line 1", "Address Line 2", "City", "State", 
            "Pin Code", "State Code", "Contact Number", "Email", 
            "GST Number", "PAN Number", "Registration Number"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add sample data
        sample_data = [
            "Sample Company Ltd",   # Name
            "123 Business Street",  # Address Line 1
            "Business Complex",     # Address Line 2
            "Mumbai",               # City
            "Maharashtra",          # State
            "400001",               # Pin Code
            "27",                   # State Code
            "1234567890",          # Contact Number
            "info@company.com",     # Email
            "27ABCDE1234F1Z5",     # GST Number
            "ABCDE1234F",          # PAN Number
            "CIN12345678"          # Registration Number
        ]

        for col, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col, value=value)
            cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Add instructions sheet
        ws_instructions = wb.create_sheet("Instructions", 0)
        instructions = [
            "Instructions for Company Import:",
            "1. Required columns must be present and spelled exactly as in the template.",
            "2. Name, Address Line 1, City, State, Pin Code, State Code, and Contact Number are mandatory.",
            "3. Address Line 2, Email, GST Number, PAN Number, and Registration Number are optional.",
            "4. Pin Code and State Code must be valid.",
            "5. If company name exists, it will be updated; otherwise, created.",
            "6. Do not modify the header row.",
            "7. Save as .xlsx format."
        ]

        for row, text in enumerate(instructions, 1):
            ws_instructions.cell(row=row, column=1, value=text)

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info("Company Excel template generated successfully")
        return excel_buffer

    @staticmethod
    def export_companies(companies_data: List[Dict]) -> io.BytesIO:
        """Export companies data to Excel with styling"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Companies Export"

        # Define styles (same as template)
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072B2", end_color="0072B2", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')

        # Add headers
        headers = [
            "Name", "Address Line 1", "Address Line 2", "City", "State",
            "Pin Code", "State Code", "Contact Number", "Email", 
            "GST Number", "PAN Number", "Registration Number"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Add data
        for row_idx, item in enumerate(companies_data, 2):
            row_data = [
                item.get("name"),
                item.get("address1"),
                item.get("address2"),
                item.get("city"),
                item.get("state"),
                item.get("pin_code"),
                item.get("state_code"),
                item.get("contact_number"),
                item.get("email"),
                item.get("gst_number"),
                item.get("pan_number"),
                item.get("registration_number")
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.border = thin_border

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 20

        # Save to BytesIO
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        logger.info(f"Companies data exported successfully: {len(companies_data)} records")
        return excel_buffer